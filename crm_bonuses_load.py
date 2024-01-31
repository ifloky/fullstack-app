import openpyxl
import psycopg2
from credentials import db_username, db_password, db_host, db_port, db_name

# Подключение к БД
connection = psycopg2.connect(
    user=db_username,
    password=db_password,
    host=db_host,
    port=db_port,
    database=db_name
)
cursor = connection.cursor()

# Открытие файла Excel
excel_file_path = '12345.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active

# Получение названий колонок из первой строки Excel
columns = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]

# Проход по строкам файла Excel, начиная со второй строки
for row_num in range(2, sheet.max_row + 1):
    # Собираем значения из строки Excel
    values = [sheet.cell(row=row_num, column=col).value for col in range(1, sheet.max_column + 1)]

    # Преобразование None в значение None для записи в БД
    values = [None if value == 'None' else value for value in values]

    # Формирование SQL-запроса для вставки данных
    query = f"INSERT INTO public.main_gameslist ({', '.join(columns)}) VALUES ({', '.join(['%s'] * len(columns))})"

    # Выполнение SQL-запроса
    cursor.execute(query, values)

# Сохранение изменений и закрытие соединения с БД
connection.commit()
cursor.close()
connection.close()
