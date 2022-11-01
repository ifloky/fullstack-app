from credentials import cc_db_host, cc_db_port, cc_db_name, cc_db_username, cc_db_password
from datetime import timedelta
from openpyxl import load_workbook
from mysql.connector import Error, connect

import pandas as pd
import time


def get_connection():
    print('Connecting to MySQL database...')
    try:
        connection = connect(host=cc_db_host,
                             port=cc_db_port,
                             database=cc_db_name,
                             user=cc_db_username,
                             password=cc_db_password)
        return connection
    except Error as e:
        print(e)


def create_df():
    print('Creating DataFrame...')
    df = []
    try:
        connection = get_connection()
        cursor = connection.cursor()
        query = (F'''
                SELECT CallDateTime, Type, Operator, Client FROM asterisk.CallsCountAll
                WHERE TYPE = 'Исходящий' AND CallDate >= '2022-10-01'
                ''')
        cursor.execute(query)
        data = cursor.fetchall()
        df = pd.DataFrame(data, columns=[c[0] for c in cursor.description])
        cursor.close()
        connection.close()
    except Error as e:
        print(e)
    return df


def load_phone_numbers():
    print('Loading phone numbers...')
    wb = load_workbook('./book2.xlsx')
    sheet = wb['по фильтру']
    max_row = str('D') + str(sheet.max_row)
    phone_numbers = []
    for cellObj in sheet['D3':max_row]:
        for cell in cellObj:
            if cell.value is not None:
                phone_number = "+" + str(cell.value).replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                phone_numbers.append(phone_number)
            else:
                continue
    print('Phone Numbers Loaded', len(phone_numbers))
    return phone_numbers


def check_call(phone_number, df):
    for index, row in df.iterrows():
        if row['client'] == phone_number:
            check = row['client'], row['CallDateTime'].strftime("%Y-%m-%d %H:%M:%S")
            check = str(check).replace('(', '').replace(')', '').replace("'", '')
            print(check)
            return check
    print(str(phone_number) + ', ' + 'No Calls')
    return str(phone_number)+', ' + 'No Calls'


def main():
    start_job_time = time.perf_counter()
    data = []
    df = create_df()
    phone_numbers = load_phone_numbers()

    for phone_number in phone_numbers:
        data.append(check_call(phone_number, df))

    file = open('result.txt', 'w')
    for item in data:
        file.write("%s \n" % item)
    file.close()

    stop_job_time = time.perf_counter()
    working_time = stop_job_time - start_job_time

    print('\n')
    print('Загружено из базы данных:', str(len(df)), 'записей')
    print('Загружено из файла:', len(phone_numbers), 'записей')
    print("Проверено и сохранено:", len(data), "номеров")
    print("Затрачено времени:", str(timedelta(seconds=working_time)))
    print(f'Время выполнения: {time.perf_counter() - start_job_time:0.4f} seconds')


if __name__ == '__main__':
    main()
